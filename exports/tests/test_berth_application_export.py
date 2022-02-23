import io

import pytest
from django.conf import settings
from django.urls import reverse
from django.utils import translation
from freezegun import freeze_time
from openpyxl import load_workbook
from rest_framework import status

from applications.models import BerthApplication, HarborChoice
from applications.schema import BerthApplicationNode
from applications.tests.conftest import generate_berth_switch_info
from applications.tests.factories import BerthApplicationFactory
from customers.tests.factories import BoatFactory
from resources.tests.factories import BoatTypeFactory, HarborFactory

from ..xlsx_writer import BerthApplicationXlsx
from .utils import to_global_ids

EXCEL_FILE_LANG = settings.LANGUAGES[0][0]


def test_amount_of_queries(superuser_api_client, django_assert_max_num_queries):
    for _i in range(2):
        harbor_1, harbor_2 = HarborFactory.create_batch(2)
        boat = BoatFactory()
        application = BerthApplicationFactory(
            boat=boat, berth_switch=generate_berth_switch_info()
        )
        HarborChoice.objects.create(
            application=application, priority=1, harbor=harbor_1
        )
        HarborChoice.objects.create(
            application=application, priority=2, harbor=harbor_2
        )

    ids = BerthApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, BerthApplicationNode)

    with django_assert_max_num_queries(3):
        response = superuser_api_client.post(
            reverse("berth_applications_xlsx"), data={"ids": global_ids}
        )

    assert response.status_code == status.HTTP_200_OK


@freeze_time("2019-01-14T08:00:00Z")
def test_export_view_produces_an_excel(superuser_api_client):
    harbor = HarborFactory()
    boat = BoatFactory()
    application = BerthApplicationFactory(boat=boat)
    HarborChoice.objects.create(application=application, priority=1, harbor=harbor)

    ids = BerthApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, BerthApplicationNode)

    response = superuser_api_client.post(
        reverse("berth_applications_xlsx"), data={"ids": global_ids}
    )

    assert response.status_code == status.HTTP_200_OK
    xlsx_bytes = response.content
    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Berth applications" in wb.sheetnames
    xl_sheet = wb["Berth applications"]

    assert xl_sheet.cell(2, 1).value == "2019-01-14 10:00"


@freeze_time("2019-01-14T08:00:00Z")
@pytest.mark.parametrize("berth_switch", [True, False])
@pytest.mark.parametrize("customer_private", [True, False])
# Parametrised berth_switch_reason inside berth_switch
@pytest.mark.parametrize("berth_switch_info", [True, False], indirect=True)
def test_exporting_berth_applications_to_excel(
    berth_switch,
    berth_switch_info,
    customer_private,
):
    harbor_1 = HarborFactory(name="Satama 1")
    harbor_1.create_translation("fi", name="Satama 1")
    harbor_2 = HarborFactory(name="Satama 2")
    harbor_2.create_translation("fi", name="Satama 2")
    boat_type = BoatTypeFactory()
    boat_type.create_translation("fi", name="Jollapaikka")
    boat_data = {
        "boat_type": boat_type,
        "width": "2",
        "length": "3.5",
        "draught": "1",
        "weight": 20000,
        "registration_number": "B0A7",
        "name": "Vene",
        "model": "BMW S 12",
        "hull_material": "wood",
        "intended_use": "cafe",
        "is_insured": True,
        "is_inspected": True,
    }
    application_data = {
        "first_name": "Kyösti",
        "last_name": "Testaaja",
        "email": "kyosti.testaaja@example.com",
        "address": "Mariankatu 2",
        "zip_code": "00170",
        "municipality": "Helsinki",
        "phone_number": "0411234567",
        "berth_switch": berth_switch_info if berth_switch else None,
        "accessibility_required": True,
        "accept_boating_newsletter": False,
        "accept_fitness_news": False,
        "accept_library_news": False,
        "accept_other_culture_news": True,
        "renting_period": "a while",
        "rent_from": "01.02.2019",
        "rent_till": "01.03.2019",
        "agree_to_terms": True,
        "application_code": "1234567890",
    }
    if not customer_private:
        application_data["company_name"] = "ACME Inc."
        application_data["business_id"] = "123123-000"
    boat = BoatFactory(**boat_data)
    application = BerthApplicationFactory(**application_data, boat=boat)
    HarborChoice.objects.create(application=application, priority=1, harbor=harbor_1)
    HarborChoice.objects.create(application=application, priority=2, harbor=harbor_2)

    expected_berth_switch_reason = None
    expected_berth_switch_str = None
    if berth_switch:
        switch_harbor_name = (
            berth_switch_info.berth.pier.harbor.safe_translation_getter(
                "name", language_code=EXCEL_FILE_LANG
            )
        )
        expected_berth_switch_str = "{} ({}): {}".format(
            switch_harbor_name,
            berth_switch_info.berth.pier.identifier,
            berth_switch_info.berth.number,
        )
        expected_berth_switch_reason = (
            berth_switch_info.reason.title if berth_switch_info.reason else "---"
        )

    harbor_1_name = harbor_1.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )
    harbor_2_name = harbor_2.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )
    boat_type_name = boat_type.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )

    with translation.override(EXCEL_FILE_LANG):
        exporter = BerthApplicationXlsx(BerthApplication.objects.all())
        xlsx_bytes = exporter.serialize()

    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Berth applications" in wb.sheetnames
    xl_sheet = wb["Berth applications"]

    assert xl_sheet.max_column == 35

    assert xl_sheet.cell(2, 1).value == "2019-01-14 10:00"
    assert xl_sheet.cell(2, 2).value == f"1: {harbor_1_name}\n2: {harbor_2_name}"
    assert xl_sheet.cell(2, 2).alignment.wrap_text is True
    assert xl_sheet.cell(2, 3).value == expected_berth_switch_str
    assert xl_sheet.cell(2, 4).value == expected_berth_switch_reason
    assert xl_sheet.cell(2, 5).value == (None if customer_private else "ACME Inc.")
    assert xl_sheet.cell(2, 6).value == (None if customer_private else "123123-000")
    assert xl_sheet.cell(2, 7).value == "Kyösti"
    assert xl_sheet.cell(2, 8).value == "Testaaja"
    assert xl_sheet.cell(2, 9).value == "kyosti.testaaja@example.com"
    assert xl_sheet.cell(2, 10).value == "Mariankatu 2"
    assert xl_sheet.cell(2, 11).value == "00170"
    assert xl_sheet.cell(2, 12).value == "Helsinki"
    assert xl_sheet.cell(2, 13).value == "0411234567"
    assert xl_sheet.cell(2, 14).value == boat_type_name
    assert xl_sheet.cell(2, 15).value == 2.0
    assert xl_sheet.cell(2, 16).value == 3.5
    assert xl_sheet.cell(2, 17).value == 1
    assert xl_sheet.cell(2, 18).value == 20000
    assert xl_sheet.cell(2, 19).value == "B0A7"
    assert xl_sheet.cell(2, 20).value == "Vene"
    assert xl_sheet.cell(2, 21).value == "BMW S 12"
    assert xl_sheet.cell(2, 22).value == "Yes"
    assert xl_sheet.cell(2, 23).value is None
    assert xl_sheet.cell(2, 24).value is None
    assert xl_sheet.cell(2, 25).value is None
    assert xl_sheet.cell(2, 26).value == "Yes"
    assert xl_sheet.cell(2, 27).value == "wood"
    assert xl_sheet.cell(2, 28).value == "cafe"
    assert xl_sheet.cell(2, 29).value == "a while"
    assert xl_sheet.cell(2, 30).value == "01.02.2019"
    assert xl_sheet.cell(2, 31).value == "01.03.2019"
    assert xl_sheet.cell(2, 32).value == "Yes"
    assert xl_sheet.cell(2, 33).value == "Yes"
    assert xl_sheet.cell(2, 34).value == "Yes"
    assert xl_sheet.cell(2, 35).value == "1234567890"
