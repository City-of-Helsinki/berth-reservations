import io

import pytest
from django.conf import settings
from django.urls import reverse
from django.utils import translation
from freezegun import freeze_time
from openpyxl import load_workbook
from rest_framework import status

from applications.enums import ApplicationAreaType, WinterStorageMethod
from applications.models import WinterStorageApplication, WinterStorageAreaChoice
from applications.schema import WinterStorageApplicationNode
from applications.tests.factories import WinterStorageApplicationFactory
from customers.tests.factories import BoatFactory
from exports.tests.utils import to_global_ids
from exports.xlsx_writer import WinterStorageApplicationXlsx
from resources.tests.factories import BoatTypeFactory, WinterStorageAreaFactory

EXCEL_FILE_LANG = settings.LANGUAGES[0][0]


def test_amount_of_queries(superuser_api_client, django_assert_max_num_queries):
    for _i in range(2):
        winter_area_1, winter_area_2 = WinterStorageAreaFactory.create_batch(2)
        boat = BoatFactory()
        application = WinterStorageApplicationFactory(
            boat=boat, area_type=ApplicationAreaType.MARKED
        )
        WinterStorageAreaChoice.objects.create(
            application=application, priority=1, winter_storage_area=winter_area_1
        )
        WinterStorageAreaChoice.objects.create(
            application=application, priority=2, winter_storage_area=winter_area_2
        )

    ids = WinterStorageApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, WinterStorageApplicationNode)

    with django_assert_max_num_queries(2):
        response = superuser_api_client.post(
            reverse("winter_storage_applications_xlsx"), data={"ids": global_ids}
        )

    assert response.status_code == status.HTTP_200_OK


@freeze_time("2019-01-14T08:00:00Z")
def test_export_view_produces_an_excel(superuser_api_client):
    winter_area = WinterStorageAreaFactory()
    boat = BoatFactory()
    application = WinterStorageApplicationFactory(
        boat=boat, area_type=ApplicationAreaType.MARKED
    )
    WinterStorageAreaChoice.objects.create(
        application=application, priority=1, winter_storage_area=winter_area
    )

    ids = WinterStorageApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, WinterStorageApplicationNode)

    response = superuser_api_client.post(
        reverse("winter_storage_applications_xlsx"), data={"ids": global_ids}
    )

    assert response.status_code == status.HTTP_200_OK
    xlsx_bytes = response.content
    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Winter storage applications" in wb.sheetnames
    xl_sheet = wb["Winter storage applications"]

    assert xl_sheet.cell(2, 1).value == "2019-01-14 10:00"


@freeze_time("2019-01-14T08:00:00Z")
@pytest.mark.parametrize("customer_private", [True, False])
def test_exporting_winter_storage_applications_to_excel(customer_private):
    winter_area_1 = WinterStorageAreaFactory(name="Satama 1")
    winter_area_1.create_translation("fi", name="Satama 1")
    winter_area_2 = WinterStorageAreaFactory(name="Satama 2")
    winter_area_2.create_translation("fi", name="Satama 2")
    boat_type = BoatTypeFactory()
    boat_type.create_translation("fi", name="Jollapaikka")
    boat_data = {
        "boat_type": boat_type,
        "width": "2",
        "length": "3.5",
        "registration_number": "B0A7",
        "name": "Vene",
        "model": "BMW S 12",
    }
    application_data = {
        "area_type": ApplicationAreaType.MARKED,
        "first_name": "Kyösti",
        "last_name": "Testaaja",
        "email": "kyosti.testaaja@example.com",
        "address": "Mariankatu 2",
        "zip_code": "00170",
        "municipality": "Helsinki",
        "phone_number": "0411234567",
        "storage_method": WinterStorageMethod.ON_TRESTLES,
        "trailer_registration_number": "hel001",
        "accept_boating_newsletter": False,
        "accept_fitness_news": False,
        "accept_library_news": False,
        "accept_other_culture_news": True,
        "application_code": "1234567890",
    }
    if not customer_private:
        application_data["company_name"] = "ACME Inc."
        application_data["business_id"] = "123123-000"

    boat = BoatFactory(**boat_data)
    application = WinterStorageApplicationFactory(**application_data, boat=boat)
    WinterStorageAreaChoice.objects.create(
        application=application, priority=1, winter_storage_area=winter_area_1
    )
    WinterStorageAreaChoice.objects.create(
        application=application, priority=2, winter_storage_area=winter_area_2
    )

    with translation.override(EXCEL_FILE_LANG):
        exporter = WinterStorageApplicationXlsx(WinterStorageApplication.objects.all())
        xlsx_bytes = exporter.serialize()

    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Winter storage applications" in wb.sheetnames

    xl_sheet = wb["Winter storage applications"]

    winter_area_name_1 = winter_area_1.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )
    winter_area_name_2 = winter_area_2.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )
    boat_type_name = boat_type.safe_translation_getter(
        "name", language_code=EXCEL_FILE_LANG
    )
    with translation.override(EXCEL_FILE_LANG):
        # Force evaluation of gettext_lazy by using str()
        # so that the translation gets evaluated in the correct language
        winter_storage_method = str(WinterStorageMethod.ON_TRESTLES.label)

    assert xl_sheet.max_column == 24

    assert xl_sheet.cell(2, 1).value == "2019-01-14 10:00"
    assert (
        xl_sheet.cell(2, 2).value == f"1: {winter_area_name_1}\n2: {winter_area_name_2}"
    )
    assert xl_sheet.cell(2, 2).alignment.wrap_text is True
    assert xl_sheet.cell(2, 3).value == (None if customer_private else "ACME Inc.")
    assert xl_sheet.cell(2, 4).value == (None if customer_private else "123123-000")
    assert xl_sheet.cell(2, 5).value == "Kyösti"
    assert xl_sheet.cell(2, 6).value == "Testaaja"
    assert xl_sheet.cell(2, 7).value == "kyosti.testaaja@example.com"
    assert xl_sheet.cell(2, 8).value == "Mariankatu 2"
    assert xl_sheet.cell(2, 9).value == "00170"
    assert xl_sheet.cell(2, 10).value == "Helsinki"
    assert xl_sheet.cell(2, 11).value == "0411234567"
    assert xl_sheet.cell(2, 12).value == winter_storage_method
    assert xl_sheet.cell(2, 13).value == "hel001"
    assert xl_sheet.cell(2, 14).value == boat_type_name
    assert xl_sheet.cell(2, 15).value == 2.0
    assert xl_sheet.cell(2, 16).value == 3.5
    assert xl_sheet.cell(2, 17).value == "B0A7"
    assert xl_sheet.cell(2, 18).value == "Vene"
    assert xl_sheet.cell(2, 19).value == "BMW S 12"
    assert xl_sheet.cell(2, 20).value is None
    assert xl_sheet.cell(2, 21).value is None
    assert xl_sheet.cell(2, 22).value is None
    assert xl_sheet.cell(2, 23).value == "Yes"
    assert xl_sheet.cell(2, 24).value == "1234567890"

    assert exporter.filename == "winter_storage_applications-2019-01-14_10-00-00"
