import io

from django.conf import settings
from django.urls import reverse
from freezegun import freeze_time
from openpyxl import load_workbook
from rest_framework import status

from applications.enums import ApplicationAreaType
from applications.models import WinterStorageApplication, WinterStorageAreaChoice
from applications.schema import WinterStorageApplicationNode
from applications.tests.factories import WinterStorageApplicationFactory
from customers.tests.factories import BoatFactory
from exports.tests.utils import to_global_ids
from resources.tests.factories import WinterStorageAreaFactory

EXCEL_FILE_LANG = settings.LANGUAGES[0][0]


def test_amount_of_queries(superuser_api_client, django_assert_max_num_queries):
    for _i in range(2):
        winter_area_1, winter_area_2 = WinterStorageAreaFactory.create_batch(2)
        boat = BoatFactory()
        application = WinterStorageApplicationFactory(
            boat=boat, area_type=ApplicationAreaType.UNMARKED
        )
        WinterStorageAreaChoice.objects.create(
            application=application, priority=1, winter_storage_area=winter_area_1
        )
        WinterStorageAreaChoice.objects.create(
            application=application, priority=2, winter_storage_area=winter_area_2
        )

    ids = WinterStorageApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, WinterStorageApplicationNode)

    with django_assert_max_num_queries(2):
        response = superuser_api_client.post(
            reverse("unmarked_winter_storage_applications_xlsx"),
            data={"ids": global_ids},
        )

    assert response.status_code == status.HTTP_200_OK


@freeze_time("2019-01-14T08:00:00Z")
def test_export_view_produces_an_excel(superuser_api_client):
    winter_area = WinterStorageAreaFactory()
    boat = BoatFactory()
    application = WinterStorageApplicationFactory(
        boat=boat, area_type=ApplicationAreaType.UNMARKED
    )
    WinterStorageAreaChoice.objects.create(
        application=application, priority=1, winter_storage_area=winter_area
    )

    ids = WinterStorageApplication.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, WinterStorageApplicationNode)

    response = superuser_api_client.post(
        reverse("unmarked_winter_storage_applications_xlsx"), data={"ids": global_ids}
    )

    assert response.status_code == status.HTTP_200_OK
    xlsx_bytes = response.content
    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Winter storage applications" in wb.sheetnames
    xl_sheet = wb["Winter storage applications"]

    assert xl_sheet.cell(2, 1).value == "2019-01-14 10:00"
