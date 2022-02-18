import datetime
import io
from unittest.mock import patch

import pytest
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.urls import reverse
from freezegun import freeze_time
from openpyxl import load_workbook
from rest_framework import status

from berth_reservations.tests.factories import CustomerProfileFactory
from customers.enums import InvoicingType
from customers.models import CustomerProfile
from customers.schema import ProfileNode
from customers.services import HelsinkiProfileUser
from exports.tests.utils import to_global_ids
from exports.xlsx_writer import CustomerXlsx


def get_mock_data_for_profiles(profiles):
    # Parse the users received from the Profile Service
    users = {}

    for profile in profiles:
        user = HelsinkiProfileUser(
            profile.id,
            email=profile.user.email if profile.user else "user@example.com",
            first_name=profile.user.first_name if profile.user else "First",
            last_name=profile.user.last_name if profile.user else "Last",
            address="Kaivokatu 1",
            postal_code="00100",
            city="Helsinki",
            phone="0501234567",
        )
        users[user.id] = user

    return users


@patch("customers.services.profile.ProfileService.get_all_profiles")
@pytest.mark.parametrize("has_permission", [True, False])
def test_admin_credentials_are_required(
    mock_get_all_profiles, user_api_client, has_permission
):
    if has_permission:
        permission = Permission.objects.get(
            content_type=ContentType.objects.get_for_model(CustomerProfile),
            codename="view_customerprofile",
        )
        user_api_client.user.user_permissions.add(permission)
    profile = CustomerProfileFactory()
    mock_get_all_profiles.return_value = get_mock_data_for_profiles([profile])
    ids = CustomerProfile.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, ProfileNode)

    response = user_api_client.post(
        reverse("customer_xlsx"), data={"ids": global_ids, "profileToken": "token"}
    )

    if has_permission:
        assert response.status_code == status.HTTP_200_OK
    else:
        assert response.status_code == status.HTTP_403_FORBIDDEN


@patch("customers.services.profile.ProfileService.get_all_profiles")
def test_amount_of_queries(
    mock_get_all_profiles, superuser_api_client, django_assert_max_num_queries
):
    CustomerProfileFactory.create_batch(2)
    CustomerProfileFactory.create_batch(2, user=None)
    mock_get_all_profiles.return_value = get_mock_data_for_profiles(
        CustomerProfile.objects.all()
    )
    ids = CustomerProfile.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, ProfileNode)

    with django_assert_max_num_queries(2):
        response = superuser_api_client.post(
            reverse("customer_xlsx"), data={"ids": global_ids, "profileToken": "token"}
        )

    assert response.status_code == status.HTTP_200_OK


@patch("customers.services.profile.ProfileService.get_all_profiles")
def test_export_view_produces_an_excel(mock_get_all_profiles, superuser_api_client):
    CustomerProfileFactory.create_batch(2)
    mock_get_all_profiles.return_value = get_mock_data_for_profiles(
        CustomerProfile.objects.all()
    )
    ids = CustomerProfile.objects.all().values_list("id", flat=True)
    global_ids = to_global_ids(ids, ProfileNode)

    response = superuser_api_client.post(
        reverse("customer_xlsx"), data={"ids": global_ids, "profileToken": "token"}
    )

    assert response.status_code == status.HTTP_200_OK
    xlsx_bytes = response.content
    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Customers" in wb.sheetnames
    xl_sheet = wb["Customers"]
    for row_number, identifier in enumerate(ids, 2):
        assert xl_sheet.cell(row_number, 1).value == str(identifier)


@patch("customers.services.profile.ProfileService.get_all_profiles")
def test_customer_excel_fields(mock_get_all_profiles):

    with freeze_time("2020-01-01T10:00:00+02:00") as frozen_datetime:
        profile_1 = CustomerProfileFactory()
        frozen_datetime.tick()
        CustomerProfileFactory()
        frozen_datetime.tick()
        CustomerProfileFactory(user=None)

    mock_get_all_profiles.return_value = get_mock_data_for_profiles([profile_1])

    expected_datetime_format = "YYYY-MM-DD HH:MM:SS"
    exporter = CustomerXlsx(
        CustomerProfile.objects.all().order_by("created_at"),
        profile_token="token",
    )

    xlsx_bytes = exporter.serialize()

    xlsx_file = io.BytesIO(xlsx_bytes)
    wb = load_workbook(filename=xlsx_file, read_only=True)
    assert "Customers" in wb.sheetnames
    xl_sheet = wb["Customers"]

    assert xl_sheet.max_column == 14

    assert xl_sheet.cell(2, 1).value == str(profile_1.id)
    assert xl_sheet.cell(2, 2).value == profile_1.user.first_name
    assert xl_sheet.cell(2, 3).value == profile_1.user.last_name
    assert xl_sheet.cell(2, 4).value == str(
        InvoicingType(profile_1.invoicing_type).label
    )
    assert xl_sheet.cell(2, 5).value == "private"
    assert xl_sheet.cell(2, 6).value == profile_1.user.email
    assert xl_sheet.cell(2, 7).value == "0501234567"
    assert xl_sheet.cell(2, 8).value == "Kaivokatu 1"
    assert xl_sheet.cell(2, 9).value == "00100"
    assert xl_sheet.cell(2, 10).value == "Helsinki"
    assert xl_sheet.cell(2, 11).value == profile_1.comment
    assert xl_sheet.cell(2, 12).value == datetime.datetime(2020, 1, 1, 10)
    assert xl_sheet.cell(2, 12).number_format == expected_datetime_format
    assert xl_sheet.cell(2, 13).value == datetime.datetime(2020, 1, 1, 10)
    assert xl_sheet.cell(2, 13).number_format == expected_datetime_format

    # User source
    assert xl_sheet.cell(2, 14).value == "Helsinki profile"
    assert xl_sheet.cell(3, 14).value == "Local"
    assert xl_sheet.cell(4, 14).value == "Local"
